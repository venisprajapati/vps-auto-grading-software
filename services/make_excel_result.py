from openpyxl.reader.excel import load_workbook


def MakeExcelResult(results, name, date):

    id_names = ReadIdNames()

    wb = load_workbook('./result/students.xlsx')

    new_sheet_ = f'{name} {date}'
    sheet_names = wb.sheetnames

    if (new_sheet_ not in sheet_names):
        ws = wb.create_sheet(new_sheet_)

        ws['A1'] = name
        ws['B1'] = date
        ws['A2'] = 'id'
        ws['B2'] = 'obtained marks'

        for index, result in enumerate(results, start=0):
            ws[f'A{index + 3}'] = result['id']
            ws[f'B{index + 3}'] = result['obtained_total_marks']

        wb.save('./result/students.xlsx')
        wb.close()

    else:
        sheet_ = wb[new_sheet_]

        length = len(sheet_['A'])

        for index, result in enumerate(results, start=length + 1):
            sheet_[f'A{index}'] = result['id']
            sheet_[f'B{index}'] = result['obtained_total_marks']

        wb.save('./result/students.xlsx')
        wb.close()

    return id_names


def ReadIdNames():
    wb = load_workbook('./result/students.xlsx')

    sheet_ = wb['Sheet1']

    id_name_dictionary = {}

    a_len = len(sheet_['A'])

    for i in range(3, a_len + 1):
        id_name_dictionary[(sheet_[f'A{i}'].value).strip()] = (sheet_[f'B{i}'].value).strip(
        ) if (sheet_[f'B{i}'].value) is not None else (sheet_[f'A{i}'].value).strip()

    wb.close()

    return id_name_dictionary
